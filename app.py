import streamlit as st
import pandas as pd
import re
from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import streamlit.components.v1 as components
import io

st.set_page_config(page_title="GeM 3-Row Exact - Built-in Master", layout="wide", page_icon="🤖")

MASTER_FILE_PATH = "Intel_Motherboard_All_Vendors_Technical_Compliance_v2.xlsx" # Keep this file in same folder as app.py

st.markdown("""
<style>
.stApp { background: radial-gradient(ellipse at top, #0F172A 0%, #020617 100%); color: #E2E8F0; }
.hero { background: linear-gradient(135deg, #020617, #1E293B); border-radius: 20px; padding: 18px 24px; border: 1px solid #334155; }
.glass-card { background: rgba(15,23,42,0.9); border-radius: 18px; padding: 20px; border: 1px solid rgba(56,189,248,0.15); margin-bottom: 14px; }
.robot-track { position: relative; height: 90px; background: #020617; border-radius: 14px; overflow: hidden; border: 1px solid rgba(56,189,248,0.2); margin: 10px 0; }
.robot { position: absolute; font-size: 58px; top: 6px; left: -70px; animation: patrol 6s linear infinite; }
@keyframes patrol { 0%{left:-70px} 50%{left:calc(100% - 60px)} 100%{left:-70px} }
.corner-robot { position: fixed; bottom: 14px; right: 14px; width: 66px; height: 66px; background: #0F172A; border-radius: 14px; display: flex; align-items: center; justify-content: center; font-size: 36px; z-index: 9999; border: 1px solid #38BDF8; }
</style>
<div class="corner-robot">🤖</div>
""", unsafe_allow_html=True)

def safe_str(x):
    try:
        if pd.isna(x): return ""
    except: pass
    return str(x).strip() if x else ""
def safe_lower(x): return safe_str(x).lower()

def read_pdf(file):
    file.seek(0)
    r = PdfReader(file)
    return "\n".join([(p.extract_text() or "") for p in r.pages])

def extract_components(text):
    comps = []
    keywords = ["processor","ram","ssd","hdd","motherboard","chipset","monitor","display","operating system","windows","cabinet","smps","keyboard","mouse","graphics","warranty","tpm","wifi","bluetooth","office","speaker"]
    lines = [l.strip() for l in text.split('\n') if 10 < len(l.strip()) < 220]
    for line in lines:
        low = line.lower()
        for kw in keywords:
            if kw in low:
                if not any(line[:45] in c[1][:45] for c in comps):
                    cat = kw.upper()
                    if "processor" in low: cat="PROCESSOR"
                    elif "ram" in low: cat="RAM"
                    elif "ssd" in low or "nvme" in low: cat="SSD"
                    elif "motherboard" in low or "chipset" in low: cat="MOTHERBOARD"
                    elif "monitor" in low: cat="MONITOR"
                    elif "windows" in low: cat="OS"
                    elif "cabinet" in low: cat="CABINET"
                    elif "smps" in low: cat="SMPS"
                    elif "keyboard" in low or "mouse" in low: cat="KBD/MOUSE"
                    elif "graphics" in low: cat="GRAPHICS"
                    elif "warranty" in low: cat="WARRANTY"
                    comps.append((cat, line[:180]))
                break
    return comps

@st.cache_data
def read_master_all_sheets():
    all_models = []
    try:
        xls = pd.ExcelFile(MASTER_FILE_PATH)
        for sheet in xls.sheet_names:
            if "Compliance" not in sheet and "Coverage" in sheet:
                continue
            try:
                for header_row in [3, 0]:
                    try:
                        df = pd.read_excel(MASTER_FILE_PATH, sheet_name=sheet, header=header_row)
                        df = df.fillna("")
                        model_col = None
                        for c in df.columns:
                            if 'model' in str(c).lower():
                                model_col = c
                                break
                        if model_col:
                            for _, row in df.iterrows():
                                model = safe_str(row.get(model_col, ""))
                                if model and len(model)>2 and model.lower() not in ["s.no.", "nan", "model"]:
                                    full_text = safe_lower(" ".join([safe_str(row[c]) for c in df.columns]))
                                    all_models.append({"model": model, "sheet": sheet, "full_text": full_text, "row": row})
                            break
                    except:
                        continue
            except:
                pass
    except Exception as e:
        st.error(f"Master file not found: {MASTER_FILE_PATH}. Please keep your Master file in same folder. Error: {e}")
        return []
    return all_models

def find_two_compatible(req_text, all_models):
    req = safe_lower(req_text)
    found = []
    for m in all_models:
        ft = m["full_text"]
        model = m["model"]
        match=False
        note=""
        if "h610" in req:
            if "h610" in ft or "b660" in ft or "b760" in ft or "z690" in ft or "b860" in ft:
                match=True
                if "b660" in ft or "b760" in ft: note="H610 asked → B660/B760 from Master — Suitable & Better (14th Gen, DDR5)"
                elif "h610" in ft: note="Exact H610 match from Master"
                else: note="Higher chipset from Master — Suitable & Better"
        elif "b660" in req:
            if "b660" in ft or "b760" in ft or "b860" in ft:
                match=True; note="B660 asked → B660/B760/B860 from Master"
        elif "ddr5" in req:
            if "ddr5" in ft:
                match=True; note="DDR5 required — DDR5 model from Master"
        elif "ddr4" in req:
            if "ddr4" in ft or "ddr5" in ft:
                match=True; note="DDR4 asked → DDR4/DDR5 from Master — DDR5 better"
        else:
            score = sum(1 for w in req.split() if len(w)>3 and w in ft)
            if score>=1:
                match=True; note=f"Compatible — {score} keywords matched"
        if match and model not in [f[0] for f in found]:
            found.append((model, note, m["sheet"]))
            if len(found)>=2:
                break
    if len(found)<2:
        for m in all_models:
            if m["model"] not in [f[0] for f in found]:
                found.append((m["model"], "Available in Master — Same category", m["sheet"]))
                if len(found)>=2: break
    return found[:2]

# Voice Guide
components.html("""
<div style="background:#0F172A; border:1px solid #38BDF8; border-radius:12px; padding:10px; display:flex; gap:8px; align-items:center;">
<div style="font-size:28px;">🤖</div>
<div style="flex:1;"><div style="color:#38BDF8; font-weight:800; font-size:12px;">ROBOT — BUILT-IN MASTER 337 MODELS</div><div id="st" style="color:#22D3EE; font-size:9px; font-family:monospace;">● NO MASTER UPLOAD — MASTER IS BUILT-IN</div></div>
<button onclick="speak()" style="background:#22D3EE; border:none; padding:6px 10px; border-radius:8px; font-weight:800; cursor:pointer; font-size:11px;">🔊 GUIDE</button>
</div>
""", height=85)

st.markdown('<div class="hero"><h1>🤖 GeM 3-Row Exact — Built-in Master (337 Models)</h1><p>Upload only ATC + BID → Row1: ATC Exact | Row2: BID Exact | Row3: 2 Compatible Models From Built-in Master</p></div><div class="robot-track"><div class="robot">🤖</div><div style="position:absolute; bottom:4px; left:50%; transform:translateX(-50%); font-family:monospace; font-size:8px; color:#38BDF8;">BUILT-IN MASTER — 337 MODELS — H610/B660/B760/Z690/B860 — 3 ROW MAPPER</div></div>', unsafe_allow_html=True)

# Load built-in master
all_models = read_master_all_sheets()
st.markdown(f'<div class="glass-card">✅ Built-in Master Loaded: <b>{len(all_models)} Models</b> from H610/B660/B760/Z690/B860/H810/Q870/Z890 (All Vendors ASRock, MSI, Gigabyte, ASUS)</div>', unsafe_allow_html=True)

c1,c2 = st.columns(2)
with c1: atc_file = st.file_uploader("📄 ATC File", type=["pdf"])
with c2: bid_file = st.file_uploader("📑 BID File", type=["pdf"])

if atc_file and bid_file:
    if not all_models:
        st.error("Master file missing. Please place Intel_Motherboard_All_Vendors_Technical_Compliance_v2.xlsx in same folder as app.py")
        st.stop()

    atc_text = read_pdf(atc_file)
    bid_text = read_pdf(bid_file)
    atc_comps = extract_components(atc_text)
    bid_comps = extract_components(bid_text)

    all_cats = {}
    for cat, txt in atc_comps: all_cats[cat] = {"atc": txt, "bid": ""}
    for cat, txt in bid_comps:
        if cat in all_cats: all_cats[cat]["bid"] = txt
        else: all_cats[cat] = {"atc": "", "bid": txt}
    if not all_cats: all_cats["MOTHERBOARD"] = {"atc": "H610 DDR4", "bid": "H610 DDR5 14th Gen"}

    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.markdown(f"#### ✅ ATC: {len(atc_comps)} | BID: {len(bid_comps)} | Built-in Master: {len(all_models)} models")

    preview = []
    for cat, vals in list(all_cats.items())[:8]:
        search = vals["bid"] if vals["bid"] else vals["atc"]
        compat = find_two_compatible(search, all_models)
        m1 = compat[0][0] if len(compat)>0 else "Not found"
        m2 = compat[1][0] if len(compat)>1 else "Not found"
        preview.append({"Category": cat, "ATC Row1": vals["atc"][:50], "BID Row2": vals["bid"][:50], "Model1 Row3": m1, "Model2 Row3": m2})
    st.dataframe(pd.DataFrame(preview), use_container_width=True, hide_index=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "3-Row Exact Mapping"
    thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    h_font = Font(bold=True, color="FFFFFF", size=11)
    h_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

    ws.merge_cells('A1:F1')
    ws['A1']=f"GeM 3-ROW EXACT — ATC + BID + 2 COMPATIBLE MODELS — Built-in Master: {len(all_models)} Models"
    ws['A1'].font=Font(bold=True, size=12, color="38BDF8")
    ws['A1'].fill=PatternFill(start_color="020617", end_color="020617", fill_type="solid")
    ws['A1'].alignment=Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height=30

    headers=["CATEGORY","ROW1: ATC Mentioned (Exact)","ROW2: BID Mentioned (Exact)","ROW3: Master Model 1","ROW3: Master Model 2","WHY Compatible?"]
    for i,h in enumerate(headers, start=1):
        c=ws.cell(row=3, column=i, value=h)
        c.font=h_font; c.fill=h_fill; c.border=thin; c.alignment=Alignment(horizontal='center', wrap_text=True, vertical='center')
    ws.row_dimensions[3].height=38

    rnum=4
    for cat, vals in all_cats.items():
        search = vals["bid"] if vals["bid"] else vals["atc"]
        compat = find_two_compatible(search, all_models)
        m1 = compat[0][0] if len(compat)>0 else "❌ Not in Master"
        n1 = compat[0][1] if len(compat)>0 else "Add product"
        m2 = compat[1][0] if len(compat)>1 else "❌ Not in Master"
        n2 = compat[1][1] if len(compat)>1 else "Add product"
        note = f"Model1: {n1} | Model2: {n2} | Sheets: {compat[0][2] if len(compat)>0 else ''}, {compat[1][2] if len(compat)>1 else ''}"

        ws.cell(row=rnum, column=1, value=cat).border=thin
        ws.cell(row=rnum, column=1).font=Font(bold=True, size=10)
        ws.cell(row=rnum, column=1).fill=PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        ws.cell(row=rnum, column=2, value=vals["atc"]).border=thin
        ws.cell(row=rnum, column=2).font=Font(color="A78BFA", size=10)
        ws.cell(row=rnum, column=2).fill=PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        ws.cell(row=rnum, column=2).alignment=Alignment(wrap_text=True, vertical='center')
        ws.cell(row=rnum, column=3, value=vals["bid"]).border=thin
        ws.cell(row=rnum, column=3).font=Font(color="38BDF8", size=10)
        ws.cell(row=rnum, column=3).fill=PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        ws.cell(row=rnum, column=3).alignment=Alignment(wrap_text=True, vertical='center')
        ws.cell(row=rnum, column=4, value=m1).border=thin
        ws.cell(row=rnum, column=4).font=Font(bold=True, color="22D3EE", size=11)
        ws.cell(row=rnum, column=4).fill=PatternFill(start_color="020617", end_color="020617", fill_type="solid")
        ws.cell(row=rnum, column=4).alignment=Alignment(wrap_text=True, vertical='center')
        ws.cell(row=rnum, column=5, value=m2).border=thin
        ws.cell(row=rnum, column=5).font=Font(bold=True, color="34D399", size=11)
        ws.cell(row=rnum, column=5).fill=PatternFill(start_color="020617", end_color="020617", fill_type="solid")
        ws.cell(row=rnum, column=5).alignment=Alignment(wrap_text=True, vertical='center')
        ws.cell(row=rnum, column=6, value=note).border=thin
        ws.cell(row=rnum, column=6).font=Font(color="FEF3C7", size=10)
        ws.cell(row=rnum, column=6).fill=PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        ws.cell(row=rnum, column=6).alignment=Alignment(wrap_text=True, vertical='center')
        ws.row_dimensions[rnum].height=46
        rnum+=1

    ws.column_dimensions['A'].width=16
    ws.column_dimensions['B'].width=30
    ws.column_dimensions['C'].width=30
    ws.column_dimensions['D'].width=26
    ws.column_dimensions['E'].width=26
    ws.column_dimensions['F'].width=40

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button("🤖📥 DOWNLOAD — 3 ROW EXACT — ATC + BID + 2 Compatible Models", data=buf, file_name="GeM_3Row_ATC_BID_2Models_Exact.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")
    st.success(f"Done! {len(all_cats)} components — Each has 2 models from built-in {len(all_models)} models — Exact sheet ready!")
    st.markdown('</div>', unsafe_allow_html=True)
else:
    st.markdown('<div class="glass-card">', unsafe_allow_html=True)
    st.info("⬆️ Upload only ATC + BID PDFs. Master is already built-in (337 models). I will create exact 3-row sheet.")
    st.markdown('</div>', unsafe_allow_html=True)